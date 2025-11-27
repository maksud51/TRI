"""
Data export utilities (JSON, CSV, Excel)
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Optional
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class DataExporter:
    """Export scraped data in multiple formats"""
    
    def __init__(self, export_path: str = 'data/exports'):
        self.export_path = Path(export_path)
        self.export_path.mkdir(parents=True, exist_ok=True)
    
    def export_json(self, profiles: List[Dict], filename: str = 'profiles.json') -> bool:
        """Export to JSON"""
        try:
            filepath = self.export_path / filename
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(profiles, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Exported {len(profiles)} profiles to {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting JSON: {e}")
            return False
    
    def export_csv(self, profiles: List[Dict], filename: str = 'profiles.csv') -> bool:
        """Export to CSV"""
        try:
            if not profiles:
                logger.warning("No profiles to export")
                return False
            
            filepath = self.export_path / filename
            
            # Flatten nested structures
            flat_profiles = []
            for profile in profiles:
                flat = self._flatten_profile(profile)
                flat_profiles.append(flat)
            
            # Get all unique keys
            all_keys = set()
            for profile in flat_profiles:
                all_keys.update(profile.keys())
            
            all_keys = sorted(list(all_keys))
            
            # Write CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=all_keys)
                writer.writeheader()
                writer.writerows(flat_profiles)
            
            logger.info(f"Exported {len(profiles)} profiles to {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return False
    
    def export_excel(self, profiles: List[Dict], filename: str = 'profiles.xlsx') -> bool:
        """Export to Excel"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not installed. Install with: pip install openpyxl")
            return False
        
        try:
            filepath = self.export_path / filename
            
            # Flatten profiles
            flat_profiles = []
            for profile in profiles:
                flat = self._flatten_profile(profile)
                flat_profiles.append(flat)
            
            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Profiles sheet
            self._create_profiles_sheet(wb, flat_profiles)
            
            # Statistics sheet
            self._create_stats_sheet(wb, profiles)
            
            # Save workbook
            wb.save(filepath)
            
            logger.info(f"Exported {len(profiles)} profiles to {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            return False
    
    def _flatten_profile(self, profile: Dict) -> Dict:
        """Flatten nested profile structures"""
        flat = {}
        
        # Basic fields
        flat['name'] = profile.get('name', '')
        flat['headline'] = profile.get('headline', '')
        flat['location'] = profile.get('location', '')
        flat['profile_url'] = profile.get('profile_url', '')
        flat['about'] = profile.get('about', '')[:500] if profile.get('about') else ''
        
        # Contact info fields - handle both single and multiple values
        contact_info = profile.get('contact_info', {})
        if contact_info:
            # Helper function to join lists or get single value
            def format_contact_field(val):
                if isinstance(val, list):
                    return ' | '.join([str(v) for v in val if v and v != 'N/A'])[:200] or 'N/A'
                return str(val) if val else 'N/A'
            
            flat['contact_emails'] = format_contact_field(contact_info.get('emails', 'N/A'))
            flat['contact_phones'] = format_contact_field(contact_info.get('phones', 'N/A'))
            flat['contact_linkedin_urls'] = format_contact_field(contact_info.get('linkedin_urls', 'N/A'))
            flat['contact_github_urls'] = format_contact_field(contact_info.get('github_urls', 'N/A'))
            flat['contact_websites'] = format_contact_field(contact_info.get('websites', 'N/A'))
            flat['contact_twitter'] = format_contact_field(contact_info.get('twitter', 'N/A'))
            flat['contact_instagram'] = format_contact_field(contact_info.get('instagram', 'N/A'))
            flat['contact_facebook'] = format_contact_field(contact_info.get('facebook', 'N/A'))
            flat['contact_whatsapp'] = format_contact_field(contact_info.get('whatsapp', 'N/A'))
            flat['contact_telegram'] = format_contact_field(contact_info.get('telegram', 'N/A'))
            flat['contact_birthday'] = format_contact_field(contact_info.get('birthday', 'N/A'))
            flat['contact_skype'] = format_contact_field(contact_info.get('skype', 'N/A'))
            flat['contact_youtube'] = format_contact_field(contact_info.get('youtube', 'N/A'))
            flat['contact_twitter_url'] = format_contact_field(contact_info.get('twitter_url', 'N/A'))
            flat['contact_linkedin_url'] = format_contact_field(contact_info.get('linkedin_url', 'N/A'))
        else:
            # Initialize all contact fields as N/A if contact_info not available
            contact_fields = ['emails', 'phones', 'linkedin_urls', 'github_urls', 'websites', 'twitter', 'instagram', 
                            'facebook', 'whatsapp', 'telegram', 'birthday', 'skype', 'youtube', 'twitter_url', 'linkedin_url']
            for field in contact_fields:
                flat[f'contact_{field}'] = 'N/A'
        
        # Experience summary
        experience = profile.get('experience', [])
        if experience:
            titles = [e.get('title', '') for e in experience[:3]]
            flat['experience_titles'] = ' | '.join(titles)
            flat['total_experience_entries'] = len(experience)
        
        # Education summary
        education = profile.get('education', [])
        if education:
            schools = [e.get('school', '') for e in education[:2]]
            flat['education_schools'] = ' | '.join(schools)
            flat['total_education_entries'] = len(education)
        
        # Skills summary
        skills = profile.get('skills', [])
        flat['skills'] = ', '.join(skills[:10]) if skills else ''
        flat['total_skills'] = len(skills)
        
        # Certifications summary
        certs = profile.get('certifications', [])
        if certs:
            flat['total_certifications'] = len(certs)
            cert_names = [c.get('name', '') for c in certs[:3]]
            flat['certifications'] = ' | '.join(cert_names)
        
        # Languages summary
        languages = profile.get('languages', [])
        flat['languages'] = ', '.join(languages) if languages else ''
        flat['total_languages'] = len(languages)
        
        # Metadata
        flat['scraped_at'] = profile.get('scraped_at', '')
        flat['extraction_method'] = profile.get('extraction_method', 'text-based')
        
        return flat
    
    def _create_profiles_sheet(self, workbook, profiles: List[Dict]):
        """Create profiles sheet in workbook"""
        ws = workbook.create_sheet('Profiles')
        
        if not profiles:
            return
        
        # Get all keys
        all_keys = sorted(set().union(*[p.keys() for p in profiles]))
        
        # Header
        for col_idx, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row_idx, profile in enumerate(profiles, 2):
            for col_idx, key in enumerate(all_keys, 1):
                value = profile.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        for col_idx in range(1, len(all_keys) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    def _create_stats_sheet(self, workbook, profiles: List[Dict]):
        """Create statistics sheet"""
        ws = workbook.create_sheet('Statistics', 0)
        
        # Calculate statistics
        total = len(profiles)
        
        # Avoid division by zero
        if total == 0:
            stats = [
                ('Metric', 'Value'),
                ('Total Profiles', 0),
                ('No profiles to analyze', ''),
            ]
        else:
            profiles_with_about = sum(1 for p in profiles if p.get('about'))
            profiles_with_exp = sum(1 for p in profiles if p.get('experience'))
            profiles_with_education = sum(1 for p in profiles if p.get('education'))
            profiles_with_skills = sum(1 for p in profiles if p.get('skills'))
            
            avg_skills = sum(len(p.get('skills', [])) for p in profiles) / total
            avg_experience = sum(len(p.get('experience', [])) for p in profiles) / total
            
            stats = [
                ('Metric', 'Value'),
                ('Total Profiles', total),
                ('Profiles with About', f"{profiles_with_about} ({profiles_with_about/total*100:.1f}%)"),
                ('Profiles with Experience', f"{profiles_with_exp} ({profiles_with_exp/total*100:.1f}%)"),
                ('Profiles with Education', f"{profiles_with_education} ({profiles_with_education/total*100:.1f}%)"),
                ('Profiles with Skills', f"{profiles_with_skills} ({profiles_with_skills/total*100:.1f}%)"),
                ('Average Skills per Profile', f"{avg_skills:.1f}"),
                ('Average Experience Entries', f"{avg_experience:.1f}"),
            ]
        
        for row_idx, (metric, value) in enumerate(stats, 1):
            ws.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
    
    def export_all_formats(self, profiles: List[Dict]) -> Dict[str, bool]:
        """Export to all available formats"""
        results = {
            'json': self.export_json(profiles),
            'csv': self.export_csv(profiles),
            'excel': self.export_excel(profiles) if OPENPYXL_AVAILABLE else None,
        }
        
        return results
    
    def get_export_path(self) -> Path:
        """Get export directory path"""
        return self.export_path
